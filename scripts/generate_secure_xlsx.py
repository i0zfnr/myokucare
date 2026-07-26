import json, sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

source, destination = sys.argv[1], sys.argv[2]
with open(source, encoding="utf-8") as handle:
    data = json.load(handle)

book = Workbook()
sheet = book.active
sheet.title = "Report"
for row in data["rows"]:
    sheet.append(row)
for cell in sheet[1]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="B63D52")
    cell.alignment = Alignment(vertical="top")
sheet.freeze_panes = "A2"
sheet.auto_filter.ref = sheet.dimensions
for column in range(1, sheet.max_column + 1):
    width = max(len(str(sheet.cell(row, column).value or "")) for row in range(1, sheet.max_row + 1))
    sheet.column_dimensions[get_column_letter(column)].width = min(max(width + 2, 12), 45)
book.save(destination)
